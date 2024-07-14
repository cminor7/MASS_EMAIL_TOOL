# ver 1.2.2
# internal libraries
from os import path, rename, getlogin
from datetime import datetime
from pathlib import Path
from shutil import rmtree
from time import sleep, localtime, strftime
from io import BytesIO

# external libraries
import pandas as pd
import openpyxl as xl
import win32com.client as win32
import pyodbc
from SMTP_email import *


def sendSupplier(test_mode, SMTP_mode, roles_selected, cc_selected=[], message='', subject=''):
	df_contact = getContact(roles_selected)
	supplier_send_file = 'supplier_send_list.xlsx'
	df_send = pd.read_excel(supplier_send_file, sheet_name='SUPPLIER_LIST', engine='openpyxl')
	current_time = strftime("%H:%M:%S", localtime())
	
	try:
		df_send['SUPPLIER_NO'] = df_send['SUPPLIER_NO'].astype('int')
	except:
		return f'[{current_time}] ERROR: BAD SUPPLIER FORMAT, TRY PASTE AS VALUES'
		
	supplier_list = df_send['SUPPLIER_NO'].unique().tolist()
	server_success_response = open('DEVELOPER_FILES/server_success_response.txt', 'r').read()
	user_email = userEmail()
	
	files = Path('ATTACHMENT').glob('*.*')
	file_list = [path.abspath(filepath) for filepath in files]

	total_supplier = len(supplier_list)
	error_counter = 0

	wb = xl.load_workbook(supplier_send_file, read_only=False, keep_vba=False)
	ws = wb['SUPPLIER_LIST']

	attachment_logic = attachmentLogic()
	if attachment_logic == 'UNIQUE FILES': # create dictionary of filepaths
		file_list_filter = [file for file in file_list if Path(file).stem.split('_')[0].strip().isdigit()]
		file_dict = {int(Path(file).stem.split('_')[0]):file for file in file_list_filter}
	
	for supplier in supplier_list:
		row_number = df_send[df_send['SUPPLIER_NO'] == supplier].index[0] + 2
				
		if ws[f'B{row_number}'].value == 'SENT' and not test_mode: # ignore sent rows in live mode
			total_supplier -= 1 
			continue

		if not test_mode: # clear out fields for new values
			for column in ws[f'B{row_number}:J{row_number}']:
				for cell in column: cell.value = ''

		ws[f'E{row_number}'] = ';'.join(roles_selected)

		if supplier not in df_contact['SUPPLIER_NUMBER'].values:
			ws[f'G{row_number}'] = 'MISSING CONTACT FIELD(S)'
			ws[f'B{row_number}'] = 'NOT SENT'
			error_counter += 1
			continue

		if attachment_logic == 'UNIQUE FILES' and supplier not in file_dict:
			ws[f'H{row_number}'] = 'MISSING FILE'
			ws[f'B{row_number}'] = 'NOT SENT'
			error_counter += 1
			continue

		current_supplier = df_contact[df_contact['SUPPLIER_NUMBER'] == supplier]
		supplier_email_list = current_supplier['SUPPLIER_EMAIL'].unique().tolist()
		supplier_email = user_email if test_mode else ';'.join(supplier_email_list)
		supplier_name = current_supplier.iloc[0]['SUPPLIER_NAME'].strip()
		supplier_no = current_supplier.iloc[0]['SUPPLIER_NUMBER']
		SPA_ID = current_supplier.iloc[0]['SPA_ID'].strip() if not current_supplier.iloc[0]['SPA_ID'] is None else ''
		SPA_name = current_supplier.iloc[0]['SPA_NAME'].strip()
		SPA_title = current_supplier.iloc[0]['SPA_TITLE'].strip()
		SPA_email = current_supplier.iloc[0]['SPA_EMAIL'].strip()
		role_flag = current_supplier.iloc[0]['ROLE_FLAG']

		ws[f'C{row_number}'] = SPA_email
		ws[f'D{row_number}'] = ';'.join(supplier_email_list)

		info_dict = {'[SUPPLIER_NAME]':supplier_name, '[SUPPLIER_NUMBER]':supplier_no, '[SPA_NAME]':SPA_name, 
			'[SPA_TITLE]':SPA_title, '[SPA_EMAIL]':SPA_email}

		em_message = msgTranslator(em_message=message, info_dict=info_dict)
		em_subject = subjTranslator(em_subject=subject, info_dict=info_dict)

		str_cc = ''
		if 'SPA' in cc_selected: str_cc = str_cc + SPA_email + ';'
		if 'USER' in cc_selected: str_cc += user_email
		em_cc = str_cc if not test_mode else ''

		em_attachment = ''
		if attachment_logic == 'STANDARD FILES':
			em_attachment = file_list[0]
		elif attachment_logic == 'UNIQUE FILES':
			em_attachment = file_dict[supplier]

		if SMTP_mode:
			server_response = SMTP(em_from=SPA_email, 
				em_to=supplier_email,
				em_cc=em_cc,
				em_subject=em_subject,
				em_message=em_message,
				em_attachment=em_attachment
				).replace('\n', ' ')
		else:
			server_response = server_success_response
			outlook(em_to=supplier_email,
				em_cc=em_cc,
				em_subject=em_subject,
				em_message=em_message,
				em_attachment=em_attachment
				)
		
		print(f'{supplier}: {SPA_email}')
		print(server_response)

		ws[f'J{row_number}'] = server_response
		ws[f'F{row_number}'] = role_flag

		try:
			server_error = serverError(response=server_response, success_response=server_success_response)
		except:
			server_error = 'CHECK SERVER_RESPONSE' 
		if server_error != 'NO ERROR':
			ws[f'B{row_number}'] = 'NOT SENT'
			ws[f'I{row_number}'] = server_error
			error_counter += 1
			continue

		elif not test_mode:
			ws[f'B{row_number}'] = 'SENT'
		
		wb.save(supplier_send_file)
		if test_mode: break
	wb.save(supplier_send_file)
	wb.close()

	if test_mode and total_supplier == error_counter:
		error_message = f'[{current_time}] TEST EMAIL FAILED'
	elif test_mode:
		error_message = f'[{current_time}] TEST EMAIL SENT'
	else:
		email_sent = total_supplier - error_counter
		error_message = f'[{current_time}]  TOTAL EMAIL: {total_supplier}  SENT: {email_sent}  FAILED: {error_counter}'
	print('\nE-MAIL SEND COMPLETE')
	return error_message


def attachmentLogic():
	files = sorted(Path('ATTACHMENT').glob('*.*'))
	file_list = [path.abspath(filepath) for filepath in files]
	
	if not file_list:
		attachment_logic = 'NO FILES'
	elif Path(file_list[0]).stem.split('_')[0].strip().isdigit():
		attachment_logic ='UNIQUE FILES'
	else:
		attachment_logic = 'STANDARD FILES'
	return attachment_logic


def userEmail():
	# get user name - outlook must be opened for this to work
	user_id = getlogin()
	gen_py = f'C:\\Users\\{user_id}\\AppData\\Local\\Temp\\gen_py'
	if path.exists(gen_py):
   		rmtree(gen_py) # remove gen_py folder
	outlook = win32.gencache.EnsureDispatch("Outlook.Application")
	return outlook.Session.CurrentUser.AddressEntry.GetExchangeUser().PrimarySmtpAddress


def getContact(roles_selected):
	# server = 'localhost\sqlexpress' # for a named instance
	# server = 'myserver,port' # to specify an alternate port
	server, database, username, password = 'server', 'database', '', '' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)
	cursor = cnxn.cursor()
	role_filter =  ', '.join(f"'{w}'" for w in roles_selected)
	query = open('DEVELOPER_FILES/contact_query.sql', 'r').read().replace("'ROLE_PLACEHOLDER'", role_filter)
	return pd.read_sql(query, cnxn)


def msgTranslator(em_message, info_dict):
	for k, v in info_dict.items(): em_message = em_message.replace(str(k), str(v))
	em_message = em_message.replace('\n', '<br>').replace('\t', '&nbsp'*8).encode('utf-8').decode('iso-8859-16')
	em_message = f"<div style='padding: 3px; width: 120px; word-wrap: break-word;'>{em_message}</div>"
	return em_message


def subjTranslator(em_subject, info_dict):
	for k, v in info_dict.items(): em_subject = em_subject.replace(str(k), str(v))
	em_subject = em_subject.encode('utf-8').decode('iso-8859-16')
	return em_subject


def hasHandle(file_path):
	try:
	    # Attempt to open the file in read mode
	    with open(file_path, 'r+') as f:
	        pass
	    return False  # File is not in use
	except IOError:
		return True  # File is in use


def serverError(response, success_response):
	if response == success_response:
		error_response = 'NO ERROR'
	elif response == '':
		error_response = 'NO SERVER RESPONSE'
	elif 'specified string is not in the form required for an e-mail address.' in response:
		error_response = 'INCORRECT E-MAIL ADDRESSES'
	elif 'Exceeded storage allocation' in response:
		error_response = 'EXCEED SIZE LIMIT'
	elif 'Unable to write data to the transport connection' in response:
		error_message = 'FAILURE TO SEND'
	else:
		error_response = 'OTHER SERVER ERROR'
	return error_response